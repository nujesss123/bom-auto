#!/usr/bin/env python3
# GitHub Actions용: 저장소의 원본 BOM 엑셀/CSV -> 암호화된 index.html 자동 재생성
# - 비밀번호는 환경변수 BOM_PASSWORD (없으면 '12345')
# - 템플릿은 저장소의 기존 index.html (앱 코드는 보존하고 데이터 블록만 교체)
import os, sys, glob, json, gzip, base64, hashlib, re, csv, datetime as dt

PW = os.environ.get('BOM_PASSWORD') or '12345'
SHELL = 'index.html'

def find_raw():
    for name in ['export.xlsx','export.csv','EXPORT.xlsx','EXPORT.csv']:
        if os.path.exists(name): return name
    xs=[f for f in glob.glob('*.xlsx') if not os.path.basename(f).startswith('~$')]
    if xs: return sorted(xs)[0]
    cs=glob.glob('*.csv')
    if cs: return sorted(cs)[0]
    return None

def read_rows(path):
    if path.lower().endswith('.csv'):
        # 인코딩 자동 시도 (utf-8-sig -> cp949)
        for enc in ('utf-8-sig','cp949','utf-8'):
            try:
                with open(path, newline='', encoding=enc) as f:
                    return [r for r in csv.reader(f)]
            except UnicodeDecodeError:
                continue
        with open(path, newline='', encoding='utf-8', errors='replace') as f:
            return [r for r in csv.reader(f)]
    try:
        from python_calamine import CalamineWorkbook
        wb=CalamineWorkbook.from_path(path)
        names=wb.sheet_names
        sn='Data' if 'Data' in names else names[0]
        return wb.get_sheet_by_name(sn).to_python(skip_empty_area=False)
    except Exception:
        import openpyxl
        wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws=wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]

def build_data(rows):
    hdr=[str(h).strip() if h is not None else '' for h in rows[0]]
    ix={h:i for i,h in enumerate(hdr)}
    def col(h): return ix.get(h,-1)
    def S(v):
        if v is None: return ''
        if isinstance(v,float) and v.is_integer(): v=int(v)
        return str(v).strip()
    def G(row,h):
        i=col(h); return '' if i<0 or i>=len(row) else S(row[i])
    def fmtDate(v):
        if isinstance(v,(dt.datetime,dt.date)): return v.strftime('%Y-%m-%d')
        if isinstance(v,(int,float)) and v>20000:
            d=dt.datetime(1899,12,30)+dt.timedelta(days=v); return d.strftime('%Y-%m-%d')
        s=str(v or '').strip()
        m=re.search(r'(\d{4})[.\-/]?(\d{2})[.\-/]?(\d{2})', s)
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else s[:10]
    def num(v):
        if v is None or v=='' : return 1
        try:
            f=float(str(v).replace(',','')); return int(f) if f.is_integer() else f
        except: return 1
    REQ=['제품코드','대체번호','제품내역(KO)','레벨번호','품목상태','재고등급','자재 유형','자재','자재 내역','소요량','기준수량','기본 단위','조달구분','자재그룹명','H자재그룹명','H제품계층구조명','브랜드명','제조사명','기종명','색상명','생성일']
    miss=[h for h in REQ if col(h)<0]
    if miss: raise SystemExit('ERROR 필수 컬럼 누락: '+', '.join(miss))
    i_code,i_ver=col('제품코드'),col('대체번호')
    i_lvl,i_qty,i_base=col('레벨번호'),col('소요량'),col('기준수량')
    i_gen,i_mc=col('생성일'),col('자재')
    mats=[];mIdx={};skus=[];sIdx={};comps=[];dates=[];n=0
    for row in rows[1:]:
        code=S(row[i_code]) if i_code>=0 and i_code<len(row) else ''
        if not code: continue
        n+=1
        ver=G(row,'대체번호')
        skey=code+'|'+ver
        if skey not in sIdx:
            ds=fmtDate(row[i_gen]) if 0<=i_gen<len(row) else ''
            if ds: dates.append(ds)
            sIdx[skey]=len(skus)
            skus.append([code,ver,G(row,'제품내역(KO)'),G(row,'제품내역(EN)'),G(row,'브랜드명'),G(row,'제조사명'),G(row,'기종명'),G(row,'색상명'),G(row,'모델'),ds,G(row,'H제품계층구조명'),G(row,'H자재그룹명'),G(row,'H MRP 관리자명'),G(row,'초안작성자')])
            comps.append([])
        si=sIdx[skey]
        mc=G(row,'자재')
        if mc not in mIdx:
            mIdx[mc]=len(mats)
            mats.append([mc,G(row,'자재 내역'),G(row,'자재그룹명'),G(row,'자재 유형'),G(row,'기본 단위')])
        qty=num(row[i_qty]) if 0<=i_qty<len(row) else 1
        base=num(row[i_base]) if 0<=i_base<len(row) else 1
        comps[si].append([G(row,'레벨번호'),mIdx[mc],qty,base,G(row,'조달구분'),G(row,'품목상태'),G(row,'재고등급'),G(row,'배치')])
    dates.sort()
    rng=f"{dates[0]} ~ {dates[-1]} 생성 BOM" if dates else '-'
    return {'mats':mats,'skus':skus,'comps':comps,'range':rng}, n

def encrypt_block(data):
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    js=json.dumps(data,ensure_ascii=False,separators=(',',':')).encode('utf-8')
    salt=os.urandom(16); iv=os.urandom(12); ITER=200000
    key=hashlib.pbkdf2_hmac('sha256',PW.encode('utf-8'),salt,ITER,dklen=32)
    ct=AESGCM(key).encrypt(iv,gzip.compress(js,9),None)
    enc=lambda x: base64.b64encode(x).decode()
    return json.dumps({'salt':enc(salt),'iv':enc(iv),'iter':ITER,'ct':enc(ct)})

def main():
    raw=find_raw()
    if not raw: raise SystemExit('ERROR 저장소에 .xlsx 또는 .csv 원본 파일이 없습니다. (예: export.xlsx 업로드)')
    if not os.path.exists(SHELL): raise SystemExit('ERROR index.html(앱 템플릿)이 저장소에 없습니다.')
    rows=read_rows(raw)
    data,nrows=build_data(rows)
    block=encrypt_block(data)
    html=open(SHELL,encoding='utf-8').read()
    html2,c=re.subn(r'(<script id="bomdata_enc" type="application/json">).*?(</script>)',
                    lambda m:m.group(1)+block+m.group(2), html, count=1, flags=re.S)
    if c!=1: raise SystemExit('ERROR index.html에서 데이터 블록(bomdata_enc)을 찾지 못했습니다.')
    open(SHELL,'w',encoding='utf-8').write(html2)
    print(f"OK '{raw}' -> index.html 재생성. 행 {nrows:,} · SKU+버전 {len(data['skus']):,} · 자재 {len(data['mats']):,} · {data['range']}")

if __name__=='__main__':
    main()
